import streamlit as st
import json
import csv
import PyPDF2
from openpyxl import load_workbook
from huggingface_hub import InferenceClient

# Hugging Face Inference Client
client = InferenceClient(
    model="microsoft/Phi-3.5-mini-instruct",
    token="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"
)

# Function to extract text from PDF
def extract_text_from_pdf(file):
    text = ""
    file.seek(0)  # Ensure the pointer is at the start of the file
    reader = PyPDF2.PdfReader(file)
    for page_num in range(len(reader.pages)):
        page = reader.pages[page_num]
        text += page.extract_text() or ""  # Extract text from each page
    return text

# Function to extract text from CSV
def extract_text_from_csv(file):
    text = ""
    file.seek(0)  # Move to the start of the file
    reader = csv.reader(file.read().decode('utf-8').splitlines())
    for row in reader:
        text += " ".join(row) + "\n"
    return text

# Function to extract text from XLSX
def extract_text_from_xlsx(file):
    text = ""
    file.seek(0)  # Ensure the pointer is at the start of the file
    workbook = load_workbook(filename=file, data_only=True)  # Read directly from file-like object
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        text += f"\nSheet: {sheet}\n"
        for row in worksheet.iter_rows(values_only=True):
            text += " ".join([str(cell) if cell is not None else "" for cell in row]) + "\n"
    return text

# Function to process the JSON file and extract data
def extract_test_report_data(json_data):
    test_cases = json_data.get("testCases", [])
    
    report_text = "Test Report Summary:\n"
    for case in test_cases:
        report_text += f"Test Name: {case['name']}\n"
        report_text += f"Status: {case['status']}\n"
        report_text += f"Execution Time: {case['executionTime']}\n"
        report_text += f"Logs: {case['logs']}\n\n"
    
    return report_text

def process_json_for_chatbot(uploaded_json):
    try:
        json_data = json.load(uploaded_json)
        return extract_test_report_data(json_data)
    except Exception as e:
        return f"Error processing JSON data: {e}"

# Function to truncate text if it exceeds a certain length
def truncate_text(text, max_length=6000):
    if len(text) > max_length:
        return text[:max_length] + "\n... [Truncated]"
    return text

# Chatbot function to generate answers based on document content
def ask_question_based_on_document(context_messages, document_text, question):
    context_messages.append({"role": "user", "content": question})
    context = (
        f"Based on the provided document, answer the following question.\n\n"
        f"Document Content:\n{truncate_text(document_text)}\n\n"
        f"Conversation history:\n"
        f"{' '.join([f'{msg['role']}: {msg['content']}' for msg in context_messages])}\n\n"
    )
    response = client.chat_completion(
        messages=[{"role": "user", "content": context}],
        max_tokens=300,
        stream=False,
    )
    context_messages.append({"role": "assistant", "content": response.choices[0].message.content})
    return response.choices[0].message.content

def ask_question_without_document(context_messages, question):
    context_messages.append({"role": "user", "content": question})
    context = (
        f"Conversation history:\n"
        f"{' '.join([f'{msg['role']}: {msg['content']}' for msg in context_messages])}\n\n"
        f"Answer the user's next question."
    )
    response = client.chat_completion(
        messages=[{"role": "user", "content": context}],
        max_tokens=300,
        stream=False,
    )
    context_messages.append({"role": "assistant", "content": response.choices[0].message.content})
    return response.choices[0].message.content

# Streamlit UI
st.title("Test Report Chatbot Assistant")
st.write("Upload a test report (PDF, CSV, XLSX, or JSON) and ask questions about it.")

# Initialize session state for chat history and document text
if "context_messages" not in st.session_state:
    st.session_state.context_messages = []
if "document_text" not in st.session_state:
    st.session_state.document_text = None

# Sidebar file upload options
uploaded_file = st.sidebar.file_uploader("Upload a file (PDF, CSV, XLSX, JSON)", type=["pdf", "csv", "xlsx", "json"])

if uploaded_file is not None:
    # Extract document content based on file type
    if uploaded_file.name.endswith(".pdf"):
        st.session_state.document_text = extract_text_from_pdf(uploaded_file)
    elif uploaded_file.name.endswith(".csv"):
        st.session_state.document_text = extract_text_from_csv(uploaded_file)
    elif uploaded_file.name.endswith(".xlsx"):
        st.session_state.document_text = extract_text_from_xlsx(uploaded_file)
    elif uploaded_file.name.endswith(".json"):
        st.session_state.document_text = process_json_for_chatbot(uploaded_file)

    st.sidebar.success("File loaded! You can now ask questions about it.")

# Display chat interface
st.write("### Ask a question:")

if "user_question" not in st.session_state:
    st.session_state.user_question = ""

user_question = st.text_input("Your question:", key="user_question_input", value=st.session_state.user_question)

# Process and respond to user question
if st.button("Send") and user_question:
    if st.session_state.document_text:
        answer = ask_question_based_on_document(st.session_state.context_messages, st.session_state.document_text, user_question)
    else:
        answer = ask_question_without_document(st.session_state.context_messages, user_question)

    st.write(f"**Assistant:** {answer}")

    # Add the user's and assistant's messages to the chat history
    st.session_state.context_messages.append({"role": "user", "content": user_question})
    st.session_state.context_messages.append({"role": "assistant", "content": answer})

    # Clear the user question input field
    st.session_state.user_question = ""  # Clear input field after sending
